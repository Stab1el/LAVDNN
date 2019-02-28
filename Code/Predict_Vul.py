# -*- coding: utf-8 -*-



from __future__ import print_function
from keras.models import Sequential
from keras import layers
import numpy as np
from six.moves import range
import sys
import os
from keras.models import load_model
import keras
import os
import time
import openpyxl
from openpyxl.workbook import *
import datetime
from vectorization import *


# Parameters for the model and dataset.
DIGITS = 50
InputChar = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ_1234567890.- '



##--------------------------read data ------------------------

f= open('final_test\ExtraFFmpeg.txt','r')
#f= open('LIBTIFF.txt','r')

data=f.read()
rows=data.split('\n')
full_data_2=[]
for row in rows:
    data=row.replace(';','')
    full_data_2.append(data)

# ---------------------add items to array------------------------
no_questions=[]
seen_2=set()
for q in full_data_2:
    if q in seen_2:
        continue
    seen_2.add(q)
    query = q + ' ' * (DIGITS - len(q))
    no_questions.append(query)

##-----------------Vectorization ----------------------------------
ctableInput = CharacterTable(InputChar)
vec_word = np.zeros((len(no_questions), DIGITS, len(InputChar)), dtype=np.bool)
for j, sentence in enumerate(no_questions):
    vec_word[j] = ctableInput.encode(sentence, DIGITS)

##----------------Prediction with model --------------------------------
model_info='F:\LSTM_TEST\model\good_model\model_2018_12_27_round_50'
model_info='F:\LSTM_TEST\model\good_model\model_2018_12_27_round_49'
model = load_model(model_info)
print('Predict model...')
print (model_info)
#preds_class_no=model.predict_classes(vec_word)
preds=model.predict(vec_word)





## ----------------- output -----------------------------------
outwb = Workbook()
wo = outwb.active
careerSheet = outwb.create_sheet('record',0)

for i in range(1,len(preds)):
    a="{0:.10f}".format(preds[i-1][0])
    careerSheet.cell(row=i,column=1).value = a
    careerSheet.cell(row=i,column=2).value = no_questions[i-1]

outwb.save("ExtraFFmpeg_test.xlsx")

'''
count_num=0
for i in range(1,len(preds)):
    if preds[i][0]>0.85:
        print (no_questions[i],preds[i][0])
        count_num=count_num+1
print (count_num)
'''

